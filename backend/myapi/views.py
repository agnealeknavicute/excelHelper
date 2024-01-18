from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Alignment
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from myapi.models import Income, Expense
from myapi.serializers import UpdatedIncomeSerializer
from rest_framework import serializers
from django.http import FileResponse
from django.views import View
from django.http import FileResponse
import os

class ExcelDownloadView(View):
    def get(self, request, *args, **kwargs):
        excel_file_path = 'income_expense_data.xlsx'

        # Проверяем существование файла
        if os.path.exists(excel_file_path):
            # Открываем файл с использованием openpyxl
            book = openpyxl.load_workbook(excel_file_path)

            # Удаляем лист "Sheet" (если он существует)
            sheet_name = 'Sheet'
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                book.remove(sheet)

            # Сохраняем изменения
            book.save(excel_file_path)

            # Отправляем файл обратно клиенту
            file = open(excel_file_path, 'rb')
            response = FileResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
            return response
        else:
            # Если файл не найден, возвращаем ошибку или другой ответ
            return HttpResponseNotFound('File not found')


class ExcelManager:
    @staticmethod
    def create_or_load_excel(excel_file_path, sheet_name):
        try:
            book = openpyxl.load_workbook(excel_file_path)
            if sheet_name not in book.sheetnames:
                book.create_sheet(sheet_name)
                book.save(excel_file_path)
        except FileNotFoundError:
            book = openpyxl.Workbook()
            book.save(excel_file_path)
            book.create_sheet(sheet_name)
            book.save(excel_file_path)

        # Загружаем данные с указанного листа
        sheet = book[sheet_name]
        return sheet

    @staticmethod
    def clear_existing_data(existing_sheet):
        # Удаляем существующие записи в колонках "name" и "value"
        existing_sheet.delete_rows(2, existing_sheet.max_row)

    @staticmethod
    def write_to_excel(existing_sheet, new_data, sheet_name, excel_file_path):
        # Если лист не содержит колонки "name" или "value", добавляем их
        if not existing_sheet['A1'].value or 'name' not in existing_sheet['A1'].value:
            existing_sheet.insert_cols(1)
            existing_sheet['A1'] = 'name'
        if not existing_sheet['B1'].value or 'value' not in existing_sheet['B1'].value:
            existing_sheet.insert_cols(2)
            existing_sheet['B1'] = 'value'

        # Получаем индексы столбцов "name" и "value"
        name_column_index = existing_sheet['A1'].column
        value_column_index = existing_sheet['B1'].column

        # Очищаем существующие данные
        ExcelManager.clear_existing_data(existing_sheet)

        # Записываем новые данные
        for row in new_data:
            # Разделяем объект на name и value
            name, value = row.get('name'), row.get('value')
            # Добавляем данные в лист
            existing_sheet.append([name, value])

        book = existing_sheet.parent
        book.save(excel_file_path)

class IncomeApi(viewsets.ModelViewSet):
    queryset = Income.objects.all()
    serializer_class = UpdatedIncomeSerializer
    http_method_names = ['get', 'post']
    permission_classes = [permissions.AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        data_list = serializer.validated_data.get('incomeItems', [])
        if data_list:
            new_data = [{'name': item['name'], 'value': item['value']} for item in data_list]

            excel_file_path = 'income_expense_data.xlsx'
            sheet_name = 'Incomes'

            # Create or load the Excel file
            existing_sheet = ExcelManager.create_or_load_excel(excel_file_path, sheet_name)

            # Write the updated data to the Excel file
            ExcelManager.write_to_excel(existing_sheet, new_data, sheet_name, excel_file_path)
        else:
            data_list = serializer.validated_data.get('expenseItems', [])
            new_data = [{'name': item['name'], 'value': item['value']} for item in data_list]

            excel_file_path = 'income_expense_data.xlsx'
            sheet_name = 'Expenses'

            # Create or load the Excel file
            existing_sheet = ExcelManager.create_or_load_excel(excel_file_path, sheet_name)

            # Write the updated data to the Excel file
            ExcelManager.write_to_excel(existing_sheet, new_data, sheet_name, excel_file_path)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        

    def perform_create(self, serializer):
        serializer.save()
